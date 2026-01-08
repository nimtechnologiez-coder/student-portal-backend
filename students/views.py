# ======================================================
# DJANGO IMPORTS
# ======================================================
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.messages import get_messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.utils import timezone
from django.db.models import Q, Count, Sum
from django.core.mail import send_mail
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum

from .models import Payment, Student

# ======================================================
# PYTHON STANDARD LIBRARY
# ======================================================
from datetime import datetime, timedelta, date
from calendar import monthrange
import random
import string
import json

# ======================================================
# EXTERNAL LIBRARIES
# ======================================================
import openpyxl
from openpyxl import Workbook
from dateutil.relativedelta import relativedelta

# ======================================================
# PROJECT MODELS
# ======================================================
from .models import (
    Student,
    Mentor,
    Topic,
    Course,
    Batch,
    Attendance,
    Holiday
)
# ======================================================
# CHECK SUPERUSER
# ======================================================
def is_admin(user):
    return user.is_superuser


# ======================================================
# COMMON LOGIN (Admin + Mentor)
# ======================================================
def admin_login(request):

    # Clear previous success messages
    storage = get_messages(request)
    for _ in storage:
        pass

    if request.method == "POST":

        input_id = request.POST.get("username")
        password = request.POST.get("password")

        user = None

        # Try email login
        user = User.objects.filter(email=input_id).first()

        # Try username
        if user is None:
            user = User.objects.filter(username=input_id).first()

        # Try mentor username_plain
        if user is None:
            mentor = Mentor.objects.filter(username_plain=input_id).first()
            if mentor:
                user = mentor.user

        # If user still not found
        if user is None:
            messages.error(request, "❌ Invalid Email or Username")
            return redirect("login")

        # Authenticate password
        auth_user = authenticate(
            request,
            username=user.username,
            password=password
        )

        if not auth_user:
            messages.error(request, "❌ Incorrect Password")
            return redirect("login")

        # Login success
        login(request, auth_user)

        # Redirect based on role
        if auth_user.is_superuser:
            return redirect("admin_dashboard")

        if Mentor.objects.filter(user=auth_user).exists():
            return redirect("mentor_today_topics")

        messages.error(request, "❌ You are not allowed to login.")
        return redirect("login")

    return render(request, "login.html")

# ======================================================
# LOGOUT
# ======================================================
def logout_user(request):
    logout(request)
    return redirect("login")



# ======================================================
# ADMIN DASHBOARD — TOPIC LIST
# ======================================================
@login_required(login_url="/")
@user_passes_test(is_admin)
def admin_dashboard(request):

    from datetime import datetime, date
    from django.db.models import Q
    from django.core.paginator import Paginator

    today = date.today()

    # --------------------------------------------------
    # GET FILTER VALUES
    # --------------------------------------------------
    search        = request.GET.get("search", "").strip()
    course_filter = request.GET.get("course", "").strip()
    batch_filter  = request.GET.get("batch", "").strip()
    month         = request.GET.get("month", "").strip()
    from_date_str = request.GET.get("from_date", "").strip()
    to_date_str   = request.GET.get("to_date", "").strip()

    # --------------------------------------------------
    # PARSE DATES SAFELY
    # --------------------------------------------------
    from_date = None
    to_date   = None

    try:
        if from_date_str:
            from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
        if to_date_str:
            to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()
    except:
        pass

    # --------------------------------------------------
    # ✅ BASE QUERY — ONLY TOPICS (NOT TASKS)
    # --------------------------------------------------
    topics = Topic.objects.select_related(
        "student__user",
        "student__course",
        "batch"
    ).filter(
        content_type="topic"   # ⭐⭐⭐ FIX ⭐⭐⭐
    )

    # --------------------------------------------------
    # CHECK IF ANY FILTER IS APPLIED
    # --------------------------------------------------
    filter_applied = any([
        search,
        course_filter,
        batch_filter,
        month,
        from_date,
        to_date,
    ])

    # --------------------------------------------------
    # DEFAULT → TODAY ONLY
    # --------------------------------------------------
    if not filter_applied:
        topics = topics.filter(date=today)

    # --------------------------------------------------
    # SEARCH FILTER
    # --------------------------------------------------
    if search:
        topics = topics.filter(
            Q(student__user__first_name__icontains=search) |
            Q(student__user__last_name__icontains=search) |
            Q(student__user__email__icontains=search) |
            Q(title__icontains=search)
        )

    # --------------------------------------------------
    # COURSE FILTER
    # --------------------------------------------------
    if course_filter:
        topics = topics.filter(student__course__course_name=course_filter)

    # --------------------------------------------------
    # BATCH FILTER
    # --------------------------------------------------
    if batch_filter:
        topics = topics.filter(batch__batch_name=batch_filter)

    # --------------------------------------------------
    # MONTH FILTER
    # --------------------------------------------------
    if month.isdigit():
        topics = topics.filter(date__month=int(month))

    # --------------------------------------------------
    # DATE RANGE FILTER
    # --------------------------------------------------
    if from_date and to_date:
        topics = topics.filter(date__range=(from_date, to_date))
    elif from_date:
        topics = topics.filter(date=from_date)

    # --------------------------------------------------
    # ORDERING
    # --------------------------------------------------
    topics = topics.order_by("-date", "-start_time")

    # --------------------------------------------------
    # CALCULATE DURATION
    # --------------------------------------------------
    for t in topics:
        try:
            start = datetime.strptime(str(t.start_time), "%H:%M:%S")
            end   = datetime.strptime(str(t.end_time), "%H:%M:%S")

            diff = end - start
            mins = diff.seconds // 60

            hrs = mins // 60
            rem = mins % 60
            dec = round(mins / 60, 2)

            t.duration_display = f"{hrs}h {rem}m ({dec} hrs)"
        except:
            t.duration_display = "—"

    # --------------------------------------------------
    # PAGINATION
    # --------------------------------------------------
    paginator = Paginator(topics, 10)
    page_obj  = paginator.get_page(request.GET.get("page"))

    params = request.GET.copy()
    params.pop("page", None)

    # --------------------------------------------------
    # CONTEXT
    # --------------------------------------------------
    context = {
        "topics": page_obj,
        "page_obj": page_obj,
        "courses": Course.objects.all().order_by("course_name"),
        "batches": Batch.objects.all().order_by("batch_name"),
        "query_string": params.urlencode(),
    }

    return render(request, "Dashboard.html", context)



# ======================================================
# ADD TOPIC — BATCH WISE
# ======================================================




# ======================================================
# STUDENT LIST
# ======================================================
@login_required(login_url="/")
def student_list(request):

    # Only admin allowed
    if not request.user.is_superuser:
        messages.error(request, "Only Admin can view student list.")
        return redirect("dashboard")

    students = Student.objects.select_related(
        "user", "course", "batch"
    ).order_by("-id")

    return render(request, "student_list.html", {
        "students": students
    })


# ======================================================
# ADD STUDENT
# ======================================================
@login_required(login_url="/")
def add_student(request):

    if not (request.user.is_superuser or hasattr(request.user, "mentor")):
        messages.error(request, "You are not allowed to access this page.")
        return redirect("login")

    if request.method == "POST":

        full_name    = request.POST.get("full_name")
        email        = request.POST.get("email")
        phone        = request.POST.get("phone")
        course_name  = request.POST.get("course_name")
        batch_name   = request.POST.get("batch_name")
        joining_date = request.POST.get("joining_date")
        duration     = request.POST.get("duration")
        amount       = request.POST.get("amount", 0)   # ✅ PAYMENT

        student_type = request.POST.get("student_type")
        access_type  = request.POST.get("access_type")
        profile_photo = request.FILES.get("profile_photo")

        # Email validation
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists!")
            return redirect("add_student")

        # Auto username + password
        clean_name = full_name.replace(" ", "").lower()
        username = clean_name + ''.join(random.choices(string.digits, k=3))

        while User.objects.filter(username=username).exists():
            username = clean_name + ''.join(random.choices(string.digits, k=3))

        password = ''.join(
            random.choices(string.ascii_letters + string.digits, k=8)
        )

        user = User.objects.create_user(
            username=username,
            first_name=full_name,
            email=email,
            password=password
        )

        # Course & batch
        course, _ = Course.objects.get_or_create(course_name=course_name)
        batch, _ = Batch.objects.get_or_create(batch_name=batch_name, course=course)

        # Dates
        joining_obj = datetime.strptime(joining_date, "%Y-%m-%d").date()
        valid_upto = joining_obj + relativedelta(months=10)

        # Student type
        if student_type == "full":
            category = "full_time"
            default_access = "all_access"
            zoom = True
        elif student_type == "part":
            category = "part_time"
            default_access = "video_only"
            zoom = False
        else:
            category = "authorized"
            default_access = "authorized_access"
            zoom = True

        access_map = {
            "All Access": "all_access",
            "Video Only Access": "video_only",
            "Authorized Access": "authorized_access",
        }

        access_final = access_map.get(access_type, default_access)

        # ✅ CREATE STUDENT (PAYMENT INCLUDED)
        Student.objects.create(
            user=user,
            phone=phone,
            course=course,
            batch=batch,
            category=category,
            access_type=access_final,
            is_zoom_enabled=zoom,
            joining_date=joining_obj,
            course_duration=duration,
            valid_upto=valid_upto,
            password_plain=password,
            amount=amount,
            profile_photo=profile_photo,
        )

        messages.success(request, "Student added successfully!")
        return redirect("student_list")

    return render(request, "add_student.html", {
        "courses": Course.objects.all(),
        "batches": Batch.objects.all(),
    })



# ======================================================
# EDIT STUDENT
# ======================================================
from datetime import datetime
@login_required(login_url="/")
def edit_student(request, student_id):

    is_admin  = request.user.is_superuser
    is_mentor = hasattr(request.user, "mentor")

    if not (is_admin or is_mentor):
        messages.error(request, "You are not allowed to access this page.")
        return redirect("login")

    student = get_object_or_404(Student, id=student_id)

    if request.method == "POST":

        username     = request.POST.get("username")
        full_name    = request.POST.get("full_name")
        email        = request.POST.get("email")
        phone        = request.POST.get("phone")
        course_name  = request.POST.get("course")
        batch_name   = request.POST.get("batch")
        joining_date = request.POST.get("joining_date")
        duration     = request.POST.get("duration") or ""
        valid_upto   = request.POST.get("valid_upto")
        amount       = request.POST.get("amount") or 0   # ✅ PAYMENT

        # -----------------------------
        # FIX DATE FORMATS
        # -----------------------------
        def fix_date(date_str):
            if not date_str:
                return None
            for fmt in ("%Y-%m-%d", "%b. %d, %Y", "%b %d, %Y"):
                try:
                    return datetime.strptime(date_str, fmt).date()
                except:
                    pass
            return None

        joining_date = fix_date(joining_date)
        valid_upto   = fix_date(valid_upto)

        # -----------------------------
        # UPDATE USER MODEL
        # -----------------------------
        user = student.user
        user.username   = username
        user.first_name = full_name
        user.email      = email
        user.save()

        # -----------------------------
        # COURSE & BATCH
        # -----------------------------
        course, _ = Course.objects.get_or_create(course_name=course_name)
        batch, _  = Batch.objects.get_or_create(batch_name=batch_name, course=course)

        # -----------------------------
        # UPDATE STUDENT
        # -----------------------------
        student.phone           = phone
        student.course          = course
        student.batch           = batch
        student.joining_date    = joining_date
        student.course_duration = duration
        student.valid_upto      = valid_upto
        student.amount          = amount   # ✅ SAVE PAYMENT

        # PROFILE PHOTO
        if request.FILES.get("profile_photo"):
            student.profile_photo = request.FILES.get("profile_photo")

        student.save()

        messages.success(request, "Student updated successfully!")
        return redirect("student_list")

    return redirect("student_list")


# ======================================================
# DELETE STUDENT
# ======================================================
@login_required(login_url="/")
def delete_student(request, student_id):

    is_admin  = request.user.is_superuser
    is_mentor = hasattr(request.user, "mentor")

    if not (is_admin or is_mentor):
        messages.error(request, "You are not allowed to delete students.")
        return redirect("login")

    student = get_object_or_404(Student, id=student_id)
    user = student.user
    user.delete()

    messages.success(request, "Student deleted successfully!")
    return redirect("student_list")


# ======================================================
# EXPORT TOPICS TO EXCEL
# ======================================================
def export_excel(request):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Topics"

    # Header
    ws.append([
        "Full Name",
        "Course",
        "Batch",
        "Date",
        "Start Time",
        "End Time",
        "Total Hours",
        "Title",
        "Description",
        "Trainer",
        "Status",
    ])

    topics = Topic.objects.select_related(
        "student__user", "student__course", "batch"
    ).all()

    for t in topics:

        # Calculate total hours
        try:
            start_dt = datetime.combine(t.date, t.start_time)
            end_dt   = datetime.combine(t.date, t.end_time)
            diff     = end_dt - start_dt
            total_hr = round(diff.total_seconds() / 3600, 2)
        except:
            total_hr = ""

        ws.append([
            f"{t.student.user.first_name} {t.student.user.last_name}",
            t.student.course.course_name,
            t.batch.batch_name,
            str(t.date),
            str(t.start_time),
            str(t.end_time),
            total_hr,
            t.title,
            t.description,
            t.trainer,
            t.status,
        ])

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="topics.xlsx"'

    wb.save(response)
    return response


# ======================================================
# EDIT TOPIC (ADMIN)
# ======================================================
@login_required
@user_passes_test(is_admin)
def edit_topic(request, topic_id):

    topic = get_object_or_404(
        Topic,
        id=topic_id,
        estimated_time__isnull=True   # ✅ PROTECT
    )

    if request.method == "POST":
        topic.title = request.POST.get("title")
        topic.description = request.POST.get("description")
        topic.date = request.POST.get("date")
        topic.start_time = request.POST.get("start_time")
        topic.end_time = request.POST.get("end_time")
        topic.trainer = request.POST.get("trainer")

        # ✅ SAFE DEFAULT
        topic.status = request.POST.get("status") or "pending"

        topic.zoom_link = request.POST.get("zoom_link")
        topic.save()

        messages.success(request, "Topic updated successfully")
        return redirect("admin_dashboard")



# ======================================================
# UPDATE TOPIC (MENTOR)
# ======================================================
@login_required
def update_topic(request):

    if request.method == "POST":

        topic_id = request.POST.get("topic_id")
        topic = get_object_or_404(Topic, id=topic_id)

        topic.title       = request.POST.get("title")
        topic.description = request.POST.get("description")
        topic.date        = request.POST.get("date")
        topic.start_time  = request.POST.get("start_time")
        topic.end_time    = request.POST.get("end_time")

        topic.save()
        messages.success(request, "Topic updated successfully!")
        return redirect("mentor_today_topics")

    return redirect("mentor_today_topics")


# ======================================================
# DELETE TOPIC (ADMIN)
# ======================================================
@login_required(login_url="/")
@user_passes_test(is_admin)
def delete_topic(request, topic_id):

    topic = get_object_or_404(Topic, id=topic_id)
    topic.delete()

    messages.success(request, "Topic deleted successfully!")
    return redirect("admin_dashboard")


# ======================================================
# DELETE TOPIC (MENTOR)
# ======================================================
@login_required(login_url="/")
def mentor_delete_topic(request, topic_id):

    mentor = Mentor.objects.filter(user=request.user).first()

    if not mentor:
        messages.error(request, "You are not a mentor.")
        return redirect("login")

    topic = get_object_or_404(Topic, id=topic_id, mentor=mentor)
    topic.delete()

    messages.success(request, "Topic deleted successfully!")
    return redirect("mentor_today_topics")


# ======================================================
# REPORTS
# ======================================================
def Reports(request):

    from_date     = request.GET.get("from_date")
    to_date       = request.GET.get("to_date")
    course_filter = request.GET.get("course")

    students = Student.objects.all()

    # Course Filter
    if course_filter:
        students = students.filter(course__course_name=course_filter)

    # Date Filter
    if from_date:
        students = students.filter(joining_date__gte=from_date)
    if to_date:
        students = students.filter(joining_date__lte=to_date)

    # Prepare Data
    report_data = []

    for s in students:

        downloads = Topic.objects.filter(student=s).aggregate(
            total=Sum("downloads")
        )["total"] or 0

        report_data.append({
            "student": s,
            "course": s.course.course_name if s.course else "-",
            "joining_date": s.joining_date,
            "end_date": s.end_date,
            "downloads": downloads
        })

    context = {
        "report_data": report_data,
        "courses": Course.objects.all(),
        "selected_course": course_filter,
        "from_date": from_date,
        "to_date": to_date,
    }

    return render(request, "reports.html", context)

# ======================================================
# NAVBAR
# ======================================================
def Navbar(request):
    return render(request, "navbar.html")


# ======================================================
# CREATE MENTOR
# ======================================================
def create_mentor(request):

    username = None
    password = None
    email_sent = False

    if request.method == "POST":

        full_name = request.POST.get("name")
        email     = request.POST.get("email")
        phone     = request.POST.get("phone")

        # -----------------------------------------
        # VALIDATION: Unique Email
        # -----------------------------------------
        if User.objects.filter(email=email).exists():
            return render(request, "create_mentor.html", {
                "error": "❌ This email is already registered."
            })

        # -----------------------------------------
        # VALIDATION: Unique Phone
        # -----------------------------------------
        if Mentor.objects.filter(phone=phone).exists():
            return render(request, "create_mentor.html", {
                "error": "❌ This phone number is already used."
            })

        # -----------------------------------------
        # AUTO GENERATE USERNAME
        # -----------------------------------------
        base = full_name.lower().replace(" ", "")
        number = random.randint(100, 999)
        username = f"{base}{number}"

        # -----------------------------------------
        # AUTO GENERATE PASSWORD
        # -----------------------------------------
        password = ''.join(random.choices(
            string.ascii_letters + string.digits, k=8
        ))

        # -----------------------------------------
        # CREATE DJANGO USER
        # -----------------------------------------
        user = User.objects.create_user(
            username=username,
            password=password,
            email=email,
            first_name=full_name
        )

        # -----------------------------------------
        # CREATE MENTOR PROFILE
        # -----------------------------------------
        Mentor.objects.create(
            user=user,
            phone=phone,
            username_plain=username,
            password_plain=password
        )

        # -----------------------------------------
        # SEND EMAIL CREDENTIALS
        # -----------------------------------------
        try:
            send_mail(
                subject="Your Mentor Login Credentials",
                message=f"""
Hello {full_name},

Your Mentor Login Details:

Username: {username}
Password: {password}

Regards,
NimTech
""",
                from_email="your_email@example.com",
                recipient_list=[email],
                fail_silently=False,
            )
            email_sent = True

        except Exception as e:
            print("EMAIL ERROR:", e)
            email_sent = False

        return render(request, "create_mentor.html", {
            "username": username,
            "password": password,
            "email": email,
            "email_sent": email_sent,
        })

    # GET request
    return render(request, "create_mentor.html")


# ======================================================
# MENTOR DASHBOARD
# ======================================================
def mentor_dashboard(request):

    mentors = Mentor.objects.all()

    return render(request, "mentor_dashboard.html", {
        "mentors": mentors
    })


# ======================================================
# DELETE MENTOR
# ======================================================
def delete_mentor(request, mentor_id):

    mentor = get_object_or_404(Mentor, id=mentor_id)
    user = mentor.user

    mentor.delete()
    user.delete()

    return redirect("mentor_dashboard")


# ======================================================
# UPDATE MENTOR
# ======================================================
def update_mentor(request):

    if request.method == "POST":

        mentor_id = request.POST.get("mentor_id")
        name      = request.POST.get("name")
        email     = request.POST.get("email")
        phone     = request.POST.get("phone")

        mentor = Mentor.objects.get(id=mentor_id)
        user   = mentor.user

        # Update User
        user.first_name = name
        user.email      = email
        user.save()

        # Update Mentor
        mentor.phone = phone
        mentor.save()

        return redirect("mentor_dashboard")


# ======================================================================
# MENTOR TODAY TOPICS
# ======================================================================
# ======================================================================
# MENTOR TODAY TOPICS  ✅ FIXED NAME
# ======================================================================
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone

from .models import Topic, Mentor, Course


# =========================================================
# MENTOR – TODAY TOPICS DASHBOARD
# =========================================================@login_required
@login_required
def mentor_today_topics(request):

    mentor = Mentor.objects.filter(user=request.user).first()

    if not mentor:
        return render(request, "mentor_today_topics.html", {
            "topics": [],
            "today": timezone.localdate(),
            "courses": Course.objects.all(),
            "error": "Mentor profile not found",
        })

    today = timezone.localdate()

    from_date = request.GET.get("from_date", "").strip()
    to_date   = request.GET.get("to_date", "").strip()
    course    = request.GET.get("course", "").strip()

    topics = Topic.objects.select_related(
        "student__user",
        "student__course",
        "batch"
    ).filter(mentor=mentor)

    # ✅ Detect filter usage
    filter_applied = any([from_date, to_date, course])

    # ✅ Default → TODAY only
    if not filter_applied:
        topics = topics.filter(date=today)

    # ✅ Filters
    if from_date:
        topics = topics.filter(date__gte=from_date)

    if to_date:
        topics = topics.filter(date__lte=to_date)

    if course:
        topics = topics.filter(
            student__course__course_name__iexact=course
        )

    topics = topics.order_by("-date", "-start_time")

    return render(request, "mentor_today_topics.html", {
        "topics": topics,
        "today": today,
        "courses": Course.objects.all(),
        "from_date": from_date,
        "to_date": to_date,
        "selected_course": course,
    })





@login_required
def mentor_delete_topic(request, topic_id):

    mentor = get_object_or_404(Mentor, user=request.user)

    topic = get_object_or_404(
        Topic,
        id=topic_id,
        mentor=mentor   # 🔐 SECURITY
    )

    topic.delete()
    messages.success(request, "Topic deleted successfully!")
    return redirect("mentor_today_topics")

@login_required
def update_topic(request):

    if request.method == "POST":

        mentor = get_object_or_404(Mentor, user=request.user)
        topic_id = request.POST.get("topic_id")

        topic = get_object_or_404(
            Topic,
            id=topic_id,
            mentor=mentor
        )

        topic.date = request.POST.get("date")
        topic.start_time = request.POST.get("start_time")
        topic.end_time = request.POST.get("end_time")
        topic.title = request.POST.get("title")
        topic.description = request.POST.get("description")

        topic.save()
        messages.success(request, "Topic updated successfully!")

    return redirect("mentor_today_topics")



# ======================================================================
# ATTENDANCE PAGE
# ======================================================================
@login_required
def attendance_page(request):

    students = Student.objects.select_related("course", "batch").all()

    today = timezone.localdate()
    three_days_before = today - timedelta(days=2)

    # Show last 3 days attendance
    records = Attendance.objects.filter(
        date__gte=three_days_before
    ).select_related(
        "student", "course", "batch"
    ).order_by("-date")

    # -----------------------------------------
    # ADD ATTENDANCE
    # -----------------------------------------
    if request.method == "POST":

        student_id = request.POST.get("student")
        date       = request.POST.get("date")
        status     = request.POST.get("status")
        remark     = request.POST.get("remark", "")

        student = get_object_or_404(Student, id=student_id)

        # Prevent Duplicate Entry
        if Attendance.objects.filter(student=student, date=date).exists():
            messages.error(request, f"⚠️ Attendance already marked for {student.user.first_name} on {date}.")
            return redirect("attendance_page")

        Attendance.objects.create(
            student=student,
            course=student.course,
            batch=student.batch,
            date=date,
            status=status,
            remark=remark
        )

        messages.success(request, "Attendance added successfully!")
        return redirect("attendance_page")

    return render(request, "attendance.html", {
        "students": students,
        "records": records
    })


# ======================================================================
# CALCULATE 6-MONTH ATTENDANCE
# ======================================================================
def calculate_6month_attendance():

    student_stats = []
    holidays = list(Holiday.objects.values_list("date", flat=True))
    students = Student.objects.select_related("user", "course", "batch").all()

    for s in students:

        if not s.joining_date:
            continue

        start = s.joining_date
        end   = start + relativedelta(months=6)

        working_days = 0
        current = start

        # Count working days
        while current <= end:

            weekday = current.weekday()  # Mon=0, Sun=6

            if weekday < 5 and current not in holidays:
                working_days += 1

            current += timedelta(days=1)

        # Count Present Days
        present_days = Attendance.objects.filter(
            student=s,
            status="Present",
            date__gte=start,
            date__lte=end
        ).count()

        percentage = round((present_days / working_days) * 100, 2) if working_days else 0

        student_stats.append({
            "name": s.user.first_name,
            "course": s.course.course_name if s.course else "",
            "batch": s.batch.batch_name if s.batch else "",
            "working_days": working_days,
            "present": present_days,
            "percentage": percentage,
        })

    return student_stats

# ======================================================
# ADMIN ATTENDANCE PAGE — FILTERS + VIEW + EXPORT
# ======================================================

@login_required
def admin_attendance_page(request):

    today = timezone.localdate()

    # -------------------- GET FILTER VALUES --------------------
    start_month = request.GET.get("start_month", "").strip()
    end_month   = request.GET.get("end_month", "").strip()
    month       = request.GET.get("month", "").strip()
    course_id   = request.GET.get("course", "").strip()
    batch_id    = request.GET.get("batch", "").strip()

    # -------------------- BASE QUERY --------------------
    records = Attendance.objects.select_related("student", "course", "batch")

    # -------------------- ANY FILTER APPLIED? --------------------
    filters = request.GET.copy()
    filters.pop("page", None)
    filter_applied = any(v.strip() for v in filters.values())

    # Default: show today's attendance
    if not filter_applied:
        records = records.filter(date=today)

    # ======================================================
    # SINGLE MONTH FILTER (YYYY-MM)
    # ======================================================
    if month:
        try:
            y, m = month.split("-")
            records = records.filter(date__year=y, date__month=m)
        except:
            pass

    # ======================================================
    # START MONTH + END MONTH RANGE
    # ======================================================
    if start_month and end_month:
        try:
            # Start month → 1st day
            y1, m1 = map(int, start_month.split("-"))
            start_date = date(y1, m1, 1)

            # End month → last day
            y2, m2 = map(int, end_month.split("-"))
            last_day = monthrange(y2, m2)[1]
            end_date = date(y2, m2, last_day)

            records = records.filter(date__range=(start_date, end_date))
        except:
            pass

    # Only Start Month
    elif start_month:
        try:
            y, m = map(int, start_month.split("-"))
            first = date(y, m, 1)
            last  = date(y, m, monthrange(y, m)[1])
            records = records.filter(date__range=(first, last))
        except:
            pass

    # Only End Month
    elif end_month:
        try:
            y, m = map(int, end_month.split("-"))
            first = date(y, m, 1)
            last  = date(y, m, monthrange(y, m)[1])
            records = records.filter(date__range=(first, last))
        except:
            pass

    # ======================================================
    # COURSE + BATCH FILTERS
    # ======================================================
    if course_id:
        records = records.filter(course_id=course_id)

    if batch_id:
        records = records.filter(batch_id=batch_id)

    # ======================================================
    # ORDER LIST
    # ======================================================
    records = records.order_by("-date", "-id")

    # ======================================================
    # EXPORT EXCEL
    # ======================================================
    if request.method == "POST" and request.POST.get("export") == "1":
        return export_attendance_excel(records)

    # ======================================================
    # LOAD FILTER OPTIONS
    # ======================================================
    courses = Course.objects.all()
    batches = Batch.objects.all()
    student_stats = calculate_6month_attendance()

    return render(request, "attendance_admin.html", {
        "records": records,
        "courses": courses,
        "batches": batches,
        "student_stats": student_stats,
        "start_month": start_month,
        "end_month": end_month,
        "month": month,
        "selected_course": course_id,
        "selected_batch": batch_id,
    })


# ======================================================
# EXPORT ATTENDANCE TO EXCEL
# ======================================================
def export_attendance_excel(records):

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header Row
    ws.append([
        "Date", "Student", "Course", "Batch", "Status", "Remark"
    ])

    # Add Data Rows
    for r in records:
        ws.append([
            r.date.strftime("%Y-%m-%d") if r.date else "",
            r.student.user.first_name if r.student else "",
            r.course.course_name if r.course else "",
            r.batch.batch_name if r.batch else "",
            r.status,
            r.remark or "",
        ])

    # Create response (download)
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=Attendance_Report.xlsx"

    wb.save(response)
    return response


# ======================================================
# EDIT ATTENDANCE
# ======================================================
@login_required
def edit_attendance(request, id):

    attendance = get_object_or_404(Attendance, id=id)

    if request.method == "POST":
        attendance.date   = request.POST.get("date")
        attendance.status = request.POST.get("status")
        attendance.remark = request.POST.get("remark")
        attendance.save()

        messages.success(request, "✏️ Attendance updated successfully!")
        return redirect("attendance_page")

    return redirect("attendance_page")


# ======================================================
# DELETE ATTENDANCE
# ======================================================
@login_required
def delete_attendance(request, id):

    attendance = get_object_or_404(Attendance, id=id)
    attendance.delete()

    messages.success(request, "🗑 Attendance deleted successfully!")
    return redirect("attendance_page")





# ============================================================================
# FRONTEND API SECTION (Login • Change Password • Student Topics)
# ============================================================================

# ======================================================
# PYTHON STANDARD LIBRARY
# ======================================================
import os
import json
import mimetypes
from io import BytesIO
from datetime import datetime, date, timedelta
from wsgiref.util import FileWrapper

# ======================================================
# DJANGO IMPORTS
# ======================================================
from django.http import (
    JsonResponse,
    StreamingHttpResponse,
    Http404,
    HttpResponse,
)
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.utils.timezone import localdate

# ======================================================
# PROJECT MODELS
# ======================================================
from students.models import (
    Student,
    Topic,
    TaskSubmission,
    Attendance,
)

# ======================================================
# DATE UTIL
# ======================================================
from dateutil.relativedelta import relativedelta

# ======================================================
# PDF (REPORTLAB)
# ======================================================
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)

# ======================================================
# EXCEL (OPENPYXL)
# ======================================================
import openpyxl
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side,
)

# ============================================================
# 1️⃣ STUDENT LOGIN API
# ============================================================
from django.utils.timezone import localdate
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.contrib.auth import authenticate
import json

@csrf_exempt
def student_login_api(request):

    if request.method != "POST":
        return JsonResponse({
            "status": "error",
            "message": "POST request required"
        })

    # Parse JSON safely
    try:
        data = json.loads(request.body.decode("utf-8"))
        username = data.get("username")
        password = data.get("password")
    except Exception:
        return JsonResponse({
            "status": "error",
            "message": "Invalid JSON"
        })

    # Authenticate User
    user = authenticate(username=username, password=password)

    if not user:
        return JsonResponse({
            "status": "error",
            "message": "Invalid username or password"
        })

    # Fetch Student Profile
    try:
        student = Student.objects.select_related(
            "course", "batch"
        ).get(user=user)
    except Student.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "Student profile not found"
        })

    # ==================================================
    # 🔒 EXPIRY CHECK (10 MONTHS)
    # ==================================================
    today = localdate()

    if student.valid_upto and today > student.valid_upto:
        return JsonResponse({
            "status": "expired",
            "message": "Your login validity has expired. Please contact admin."
        })

    # ==================================================
    # LOGIN SUCCESS
    # ==================================================
    return JsonResponse({
        "status": "success",
        "message": "Login successful",

        # USER DETAILS
        "user_id": user.id,
        "username": user.username,
        "name": user.first_name,
        "email": user.email,

        # STUDENT DETAILS
        "phone": student.phone,
        "course": student.course.course_name if student.course else None,
        "batch": student.batch.batch_name if student.batch else None,
        "access_type": student.access_type,
        "valid_upto": student.valid_upto.strftime("%Y-%m-%d"),

        # PROFILE PHOTO
        "profile_photo": (
            student.profile_photo.url
            if student.profile_photo else None
        )
    })



# ============================================================
# 2️⃣ CHANGE PASSWORD API
# ============================================================
@csrf_exempt
def change_password_api(request):

    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "POST request required"})

    # Parse JSON safely
    try:
        data = json.loads(request.body)
    except:
        return JsonResponse({"status": "error", "message": "Invalid JSON"})

    username = data.get("username")
    old_password = data.get("old_password")
    new_password = data.get("new_password")
    confirm_password = data.get("confirm_password")

    # Validate fields
    if not all([username, old_password, new_password, confirm_password]):
        return JsonResponse({"status": "error", "message": "All fields are required"})

    if new_password != confirm_password:
        return JsonResponse({"status": "error", "message": "Passwords do not match"})

    # Get User
    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        return JsonResponse({"status": "error", "message": "User not found"})

    # Get Student object
    try:
        student = Student.objects.get(user=user)
    except Student.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Student not found"})

    # Validate old password (Django or plain stored)
    if not user.check_password(old_password) and student.password_plain != old_password:
        return JsonResponse({"status": "error", "message": "Invalid old password"})

    # Update passwords
    user.set_password(new_password)
    user.save()

    student.password_plain = new_password
    student.save()

    return JsonResponse({
        "status": "success", 
        "message": "Password updated successfully!" 
    })


# ============================================================================
# 3️⃣ STUDENT TOPICS API (GET / POST)
@csrf_exempt
def student_topics_api(request):

    user_id = request.GET.get("user_id") or request.POST.get("user_id")

    if not user_id:
        return JsonResponse({"error": "user_id is required"}, status=400)

    student = Student.objects.filter(user__id=user_id).first()

    if not student:
        return JsonResponse({"error": "Student not found"}, status=404)

    topics = Topic.objects.filter(student=student).order_by("-date", "-start_time")

    topic_list = []
    total_hours = 0
    completed_tasks = 0
    pending_tasks = 0
    graph_points = []

    for t in topics:

        # calculate total hours
        if t.start_time and t.end_time:
            dt1 = datetime.combine(date.today(), t.start_time)
            dt2 = datetime.combine(date.today(), t.end_time)
            hours = (dt2 - dt1).seconds / 3600
        else:
            hours = float(t.total_hours or 0)

        total_hours += hours

        # task status
        if getattr(t, "status", "").lower() == "completed":
            completed_tasks += 1
        else:
            pending_tasks += 1

        topic_list.append({
            "id": t.id,
            "date": t.date,
            "start_time": str(t.start_time),
            "end_time": str(t.end_time),
            "total_hours": hours,
            "title": t.title,
            "description": t.description,
            "trainer": t.trainer,
            "video": f"/api/student/video/stream/{t.id}/" if t.video else None,
            "zoom_link": t.zoom_link if student.access_type == "all_access" else None
        })

    # graph points last 7
    last_7 = topics.order_by("date")[:7]
    for t in last_7:
        graph_points.append({
            "x": str(t.date),
            "y": float(getattr(t, "total_hours", 0) or 0)
        })

    # task completion %
    total_tasks = completed_tasks + pending_tasks
    task_percentage = round((completed_tasks / total_tasks) * 100, 2) if total_tasks else 0

    # FINAL RESPONSE
    return JsonResponse({
        "student_name": student.user.first_name,
        "course_name": student.course.course_name if student.course else "",
        "batch_name": student.batch.batch_name if student.batch else "",
        "access_type": student.access_type,

        "profile_photo": student.profile_photo.url if student.profile_photo else None,  # ⭐ NEW ⭐

        "total_hours": round(total_hours, 1),
        "completed_tasks": completed_tasks,
        "pending_tasks": pending_tasks,
        "task_percentage": task_percentage,
        "graph_points": graph_points,

        "topics": topic_list
    })




# ======================================================================
# 1️⃣ VIDEO STREAM — NO DOWNLOAD ALLOWED
# ======================================================================
def stream_video(request, topic_id):
    try:
        topic = Topic.objects.get(id=topic_id)
    except Topic.DoesNotExist:
        raise Http404("Video not found")

    video_path = topic.video.path
    if not os.path.exists(video_path):
        raise Http404("Video file missing")

    file = open(video_path, "rb")

    response = StreamingHttpResponse(
        FileWrapper(file, blksize=1024 * 64),
        content_type=mimetypes.guess_type(video_path)[0] or "video/mp4"
    )

    # ❌ Prevent download
    response["Content-Disposition"] = "inline"
    response["Accept-Ranges"] = "bytes"
    response["X-Content-Type-Options"] = "nosniff"
    response["Cache-Control"] = "no-store"

    # ❌ Removes Save-As option in most browsers
    response["Content-Security-Policy"] = "sandbox allow-scripts allow-same-origin"

    return response

# ============================================================================

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from students.models import Topic



# ---------------------------------------------------
# 🟥 DELETE TASK
# ---------------------------------------------------_login
@login_required
def delete_topic(request, topic_id):

    topic = get_object_or_404(Topic, id=topic_id)
    topic.delete()

    messages.success(request, "Task deleted successfully!")
    return redirect("task_list")


# =======================================================
# 🔹 HELPER: SAFE HOURS CALCULATION
# =======================================================
def _hours_between(start, end):
    """Returns hour difference between two times safely."""
    try:
        if not start or not end:
            return 0

        if isinstance(start, str):
            start = datetime.strptime(start, "%H:%M:%S").time()
        if isinstance(end, str):
            end = datetime.strptime(end, "%H:%M:%S").time()

        dt_start = datetime.combine(date.today(), start)
        dt_end   = datetime.combine(date.today(), end)

        if dt_end < dt_start:
            return 0

        return round((dt_end - dt_start).total_seconds() / 3600, 2)

    except:
        return 0


# =======================================================
# 🔹 1️⃣ STUDENT PROFILE API
# =======================================================
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from students.models import Student


@csrf_exempt
def student_profile_api(request):
    user_id = request.GET.get("user_id")

    if not user_id:
        return JsonResponse(
            {"status": "error", "message": "user_id required"},
            status=400
        )

    student = (
        Student.objects
        .select_related("user", "course", "batch")
        .filter(user__id=user_id)
        .first()
    )

    if not student:
        return JsonResponse(
            {"status": "error", "message": "Student not found"},
            status=404
        )

    profile_photo = (
        request.build_absolute_uri(student.profile_photo.url)
        if student.profile_photo else None
    )

    return JsonResponse({
        "status": "success",
        "profile": {
            "name": student.user.first_name or student.user.username,
            "email": student.user.email,
            "mobile": str(student.phone or ""),   # ✅ FIX
            "course": student.course.course_name if student.course else "",
            "batch": student.batch.batch_name if student.batch else "",
            "profile_photo": profile_photo
        }
    })



# =======================================================
# 🔹 2️⃣ STUDENT DASHBOARD API
# =======================================================
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now
from datetime import timedelta
from .models import Student, Topic, Attendance


@csrf_exempt
def student_dashboard_api(request):
    user_id = request.GET.get("user_id")

    if not user_id:
        return JsonResponse(
            {"status": "error", "message": "user_id required"},
            status=400
        )

    student = Student.objects.select_related(
        "user", "course", "batch"
    ).filter(user__id=user_id).first()

    if not student or not student.joining_date:
        return JsonResponse(
            {"status": "error", "message": "Student or joining date missing"},
            status=404
        )

    # ==================================================
    # FIXED 90 DAYS WINDOW
    # ==================================================
    start_date = student.joining_date
    end_date = start_date + relativedelta(months=3)
    TOTAL_DAYS = 90

    attendance_qs = Attendance.objects.filter(
        student=student,
        date__range=(start_date, end_date)
    )

    present_days = attendance_qs.filter(status="Present").count()
    late_days = attendance_qs.filter(status="Late").count()

    attended_days = present_days + late_days

    attendance_percent = round(
        (attended_days / TOTAL_DAYS) * 100
    )

    # ==================================================
    # COURSE PROGRESS (SAME 90 DAYS)
    # ==================================================
    course_percent = attendance_percent  # ✅ SAME BASIS

    # ==================================================
    # PROFILE
    # ==================================================
    profile = {
        "name": student.user.first_name,
        "email": student.user.email,
        "course": student.course.course_name if student.course else "",
        "batch": student.batch.batch_name if student.batch else "",
    }

    # ==================================================
    # TASK COMPLETION (MODULE BASED)
    # ==================================================
    completed_tasks = Topic.objects.filter(
        student=student,
        status="completed"
    ).count()

    task_percent = min(completed_tasks * 10, 100)

    # ==================================================
    # WEEKLY TASK GRAPH
    # ==================================================
    today = timezone.localdate()
    week_start = today - timedelta(days=today.weekday())
    days = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

    weekly_graph = []

    for i in range(7):
        day = week_start + timedelta(days=i)

        completed = Topic.objects.filter(
            student=student,
            date=day,
            status="completed"
        ).count()

        pending = Topic.objects.filter(
            student=student,
            date=day
        ).exclude(status="completed").count()

        weekly_graph.append({
            "day": days[i],
            "value": completed - pending
        })

    # ==================================================
    # FINAL RESPONSE
    # ==================================================
    return JsonResponse({
        "status": "success",

        "profile": profile,

        "attendance_progress": attendance_percent,
        "course_progress": course_percent,

        "task_completion_percent": task_percent,
        "weekly_graph": weekly_graph,
    })



# =======================================================
# 🔹 3️⃣ COURSE PROGRESS (ATTENDANCE BASED)
# =======================================================
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from .models import Student, Attendance, Topic,Payment, Student

@csrf_exempt
def student_course_progress_api(request):
    user_id   = request.GET.get("user_id")
    from_date = request.GET.get("from_date")
    end_date  = request.GET.get("end_date")

    if not user_id:
        return JsonResponse(
            {"status": "error", "message": "user_id required"},
            status=400
        )

    student = Student.objects.filter(user__id=user_id).first()
    if not student:
        return JsonResponse(
            {"status": "error", "message": "Student not found"},
            status=404
        )

    if not student.joining_date:
        return JsonResponse(
            {"status": "error", "message": "Joining date missing"},
            status=400
        )

    # ==================================================
    # 1️⃣ COURSE PROGRESS
    # ==================================================
    course_start = student.joining_date
    course_end   = student.joining_date + relativedelta(months=3)

    total_days = (course_end - course_start).days + 1

    attendance_all = Attendance.objects.filter(
        student=student,
        date__range=(course_start, course_end)
    )

    present_days = attendance_all.filter(status="Present").count()

    progress_percent = round((present_days / total_days) * 100, 2)
    highest_score    = min(present_days * 5, 100)

    # ==================================================
    # 2️⃣ DATE FILTER LOGIC
    # ==================================================
    today = date.today()

    if from_date and end_date:
        table_start = datetime.strptime(from_date, "%Y-%m-%d").date()
        table_end   = datetime.strptime(end_date, "%Y-%m-%d").date()

    elif from_date:
        table_start = table_end = datetime.strptime(from_date, "%Y-%m-%d").date()

    else:
        # ✅ DEFAULT → TODAY ONLY
        table_start = table_end = today

    topics_qs = Topic.objects.filter(
        student=student,
        date__range=(table_start, table_end)
    ).order_by("date", "start_time")

    # ==================================================
    # 3️⃣ TABLE DATA
    # ==================================================
    table_data = []

    for t in topics_qs:

        # Trainer name
        trainer = (
            t.trainer
            or (t.mentor.user.first_name if t.mentor else "N/A")
        )

        # Video URL
        video_url = (
            request.build_absolute_uri(
                f"/api/student/video/stream/{t.id}/"
            )
            if t.video else None
        )

        # ⏱ CALCULATE TOTAL HOURS (SAFE)
        total_hours = "—"
        if t.start_time and t.end_time:
            start = datetime.combine(date.today(), t.start_time)
            end   = datetime.combine(date.today(), t.end_time)
            diff  = end - start

            minutes = diff.seconds // 60
            hrs     = minutes // 60
            mins    = minutes % 60

            total_hours = f"{hrs}h {mins}m"

        table_data.append({
            "date": t.date.strftime("%d/%m/%Y"),
            "topic": t.title,
            "trainer": trainer,

            # ✅ NEW FIELDS (VERY IMPORTANT)
            "start_time": t.start_time.strftime("%H:%M") if t.start_time else "—",
            "end_time":   t.end_time.strftime("%H:%M") if t.end_time else "—",
            "hours":      total_hours,

            "status": t.status.capitalize(),
            "zoom_link": t.zoom_link or "N/A",
            "video": video_url,
        })

    # ==================================================
    # FINAL RESPONSE
    # ==================================================
    return JsonResponse({
        "status": "success",

        "student": student.user.first_name,
        "course": student.course.course_name if student.course else None,
        "batch": student.batch.batch_name if student.batch else None,

        "progress": progress_percent,
        "highest_score": highest_score,
        "present_days": present_days,
        "total_days": total_days,

        "from_date": table_start.strftime("%Y-%m-%d"),
        "end_date": table_end.strftime("%Y-%m-%d"),

        # ✅ FRONTEND TABLE
        "topics": table_data,
    })





@csrf_exempt
def student_attendance_dashboard_api(request):
    """
    ATTENDANCE DASHBOARD API (3 MONTH BASED)

    Calculates:
    - Total attendance (3 months)
    - Late / Absent
    - Successful Attendance % (Present + Late)
    - On-Time % / Late %
    - Weekly Attendance (Mon–Sun) [UI graph]
    """

    user_id = request.GET.get("user_id")
    if not user_id:
        return JsonResponse(
            {"status": "error", "message": "user_id required"},
            status=400
        )

    student = Student.objects.select_related("user").filter(
        user__id=user_id
    ).first()

    if not student or not student.joining_date:
        return JsonResponse(
            {"status": "error", "message": "Student or joining date missing"},
            status=404
        )

    # ==================================================
    # 📅 3 MONTH DATE RANGE
    # ==================================================
    start_date = student.joining_date
    end_date = start_date + relativedelta(months=3)

    total_days = (end_date - start_date).days + 1

    attendance_qs = Attendance.objects.filter(
        student=student,
        date__range=(start_date, end_date)
    )

    # ==================================================
    # 📊 COUNTS
    # ==================================================
    present_days = attendance_qs.filter(status="Present").count()
    late_days = attendance_qs.filter(status="Late").count()
    absent_days = attendance_qs.filter(status="Absent").count()
    leave_days = attendance_qs.filter(status="Leave").count()

    total_attendance = present_days + late_days + absent_days + leave_days

    # ==================================================
    # 📈 DAILY-WISE SUCCESS (3 MONTH)
    # ==================================================
    successful_days = present_days + late_days

    successful_percentage = round(
        (successful_days / total_days) * 100, 2
    ) if total_days else 0

    on_time_percentage = round(
        (present_days / total_days) * 100, 2
    ) if total_days else 0

    late_percentage = round(
        (late_days / total_days) * 100, 2
    ) if total_days else 0

    # ==================================================
    # 📊 WEEKLY ATTENDANCE (UI GRAPH – MON → SUN)
    # ==================================================
    today = localdate()
    start_of_week = today - timedelta(days=today.weekday())

    week_labels = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    weekly_attendance = []

    for i in range(7):
        current_date = start_of_week + timedelta(days=i)
        record = attendance_qs.filter(date=current_date).first()

        if record:
            if record.status == "Present":
                value = 1
            elif record.status == "Late":
                value = 0.5
            else:
                value = 0
            status = record.status
        else:
            value = 0
            status = "Absent"

        weekly_attendance.append({
            "day": week_labels[i],
            "date": current_date.isoformat(),
            "status": status,
            "value": value,
        })

    # ==================================================
    # ✅ FINAL RESPONSE (MATCHES REACT)
    # ==================================================
    return JsonResponse({
        "status": "success",

        "student": {
            "name": student.user.first_name or student.user.username,
            "email": student.user.email,
            "profile_photo": (
                request.build_absolute_uri(student.profile_photo.url)
                if student.profile_photo else None
            ),
        },

        "summary": {
            "total_attendance": total_attendance,
            "late_days": late_days,
            "absent_days": absent_days,
        },

        "percentages": {
            "successful_attendance": successful_percentage,
            "on_time": on_time_percentage,
            "late": late_percentage,
        },

        "weekly_attendance": weekly_attendance,
    })





@csrf_exempt
def student_task_log_api(request):
    """
    STUDENT TASK LOG API
    - GET  → Show student info + tasks
    - POST → Upload task PDF (student only)
    """

    user_id = request.GET.get("user_id") or request.POST.get("user_id")
    if not user_id:
        return JsonResponse(
            {"status": "error", "message": "user_id required"},
            status=400
        )

    student = (
        Student.objects
        .select_related("user", "course", "batch")
        .filter(user__id=user_id)
        .first()
    )

    if not student:
        return JsonResponse(
            {"status": "error", "message": "Student not found"},
            status=404
        )

    # ==================================================
    # POST → UPLOAD TASK PDF
    # ==================================================
    if request.method == "POST":
        topic_id = request.POST.get("topic_id")
        file = request.FILES.get("file")

        if not topic_id or not file:
            return JsonResponse(
                {"status": "error", "message": "topic_id and file required"},
                status=400
            )

        if not file.name.lower().endswith(".pdf"):
            return JsonResponse(
                {"status": "error", "message": "Only PDF allowed"},
                status=400
            )

        topic = get_object_or_404(
            Topic,
            id=topic_id,
            student=student,
            content_type="task"   # 🔥 ENSURE TASK ONLY
        )

        submission, created = TaskSubmission.objects.get_or_create(
            topic=topic,
            student=student,
            defaults={"file": file}
        )

        if not created:
            submission.file = file
            submission.save()

        # Mark task as review
        topic.status = "review"
        topic.save(update_fields=["status"])

        return JsonResponse({
            "status": "success",
            "message": "Task submitted successfully"
        })

    # ==================================================
    # GET → TASK LIST (🔥 FIXED)
    # ==================================================
    tasks = (
        Topic.objects
        .filter(
            student=student,
            content_type="task"   # 🔥 THIS WAS THE BUG
        )
        .select_related("mentor", "mentor__user")
        .prefetch_related("submissions")
        .order_by("-date", "-id")
    )

    task_list = []

    for t in tasks:
        submission = t.submissions.first()

        # Mentor name logic
        mentor_name = "N/A"
        if t.mentor and t.mentor.user:
            mentor_name = (
                t.mentor.user.get_full_name()
                or t.mentor.user.username
            )
        elif t.trainer:
            mentor_name = t.trainer

        task_list.append({
            "id": t.id,
            "date": t.date.strftime("%d/%m/%Y"),
            "topic": t.title,
            "mentor": mentor_name,
            "deadline": (
                t.deadline.strftime("%d/%m/%Y")
                if t.deadline else "-"
            ),
            "status": t.status,
            "submitted": bool(submission),
        })

    return JsonResponse({
        "status": "success",
        "student": {
            "name": student.user.first_name or student.user.username,
            "email": student.user.email,
            "course": student.course.course_name if student.course else "",
            "batch": student.batch.batch_name if student.batch else "",
        },
        "tasks": task_list
    })




from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from .models import Student

@csrf_exempt
def student_settings_api(request):
    """
    Student Settings API
    Returns:
    - Name
    - Email
    - Mobile Number
    - Course
    - Batch
    - Profile Photo
    """

    user_id = request.GET.get("user_id")

    if not user_id:
        return JsonResponse(
            {"status": "error", "message": "user_id required"},
            status=400
        )

    # -----------------------------
    # GET USER
    # -----------------------------
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse(
            {"status": "error", "message": "User not found"},
            status=404
        )

    # -----------------------------
    # GET STUDENT
    # -----------------------------
    student = Student.objects.filter(user=user).first()

    if not student:
        return JsonResponse(
            {"status": "error", "message": "Student not found"},
            status=404
        )

    # -----------------------------
    # RESPONSE
    # -----------------------------
    return JsonResponse({
        "status": "success",
        "profile": {
            "name": user.first_name or user.username,
            "email": user.email,
            "mobile": str(student.phone or ""),   # ✅ GUARANTEED STRING
            "course": (
                student.course.course_name
                if student.course else ""
            ),
            "batch": (
                student.batch.batch_name
                if student.batch else ""
            ),
            "profile_photo": (
                request.build_absolute_uri(student.profile_photo.url)
                if student.profile_photo else None
            )
        }
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.contrib import messages
from .models import Student, Topic



def is_admin_or_mentor(user):
    return user.is_staff or hasattr(user, "mentor")


# ==================================================
# ADMIN – PAYMENT LIST (COURSE / PAID / BALANCE)
# ==================================================
@staff_member_required
def admin_payment_list(request):
    payments = (
        Payment.objects
        .select_related(
            "student",
            "student__user",
            "student__course"
        )
        .order_by("-created_at")
    )

    # ✅ PRE-CALCULATE APPROVED PAYMENTS (FAST)
    approved_map = (
        Payment.objects
        .filter(status="approved")
        .values("student_id")
        .annotate(total=Sum("amount_paid"))
    )
    approved_dict = {x["student_id"]: float(x["total"]) for x in approved_map}

    # ✅ ATTACH VALUES PER PAYMENT ROW
    for p in payments:
        course_amount = float(
            getattr(p.student, "amount", 0) or
            getattr(p.student.course, "total_fee", 0) or
            0
        )

        approved_paid = approved_dict.get(p.student_id, 0)
        balance_amount = max(course_amount - approved_paid, 0)

        p.course_amount = course_amount
        p.approved_paid = approved_paid
        p.balance_amount = balance_amount

    return render(
        request,
        "payment_list.html",
        {"payments": payments}
    )


# ==================================================
# ADMIN – APPROVE PAYMENT
# ==================================================
@staff_member_required
def approve_payment(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)

    if payment.status != "approved":
        payment.status = "approved"
        payment.save(update_fields=["status"])

        messages.success(
            request,
            f"₹{payment.amount_paid} payment approved"
        )

    return redirect("admin_payment_list")


# ==================================================
# ADMIN – REJECT PAYMENT
# ==================================================
@staff_member_required
def reject_payment(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)

    if request.method == "POST":
        reason = request.POST.get("reason")

        if not reason:
            messages.error(request, "Rejection reason is required")
            return redirect("admin_payment_list")

        payment.status = "rejected"
        payment.admin_remark = reason
        payment.save(update_fields=["status", "admin_remark"])

        messages.error(
            request,
            f"₹{payment.amount_paid} payment rejected"
        )

    return redirect("admin_payment_list")


# ==================================================
# STUDENT – SUBMIT PAYMENT (API)
# ==================================================
@csrf_exempt
def submit_payment_api(request):
    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "message": "POST only"},
            status=400
        )

    # ✅ READ FORM DATA
    user_id = request.POST.get("user_id")
    amount = request.POST.get("amount")
    utr = request.POST.get("utr")
    screenshot = request.FILES.get("screenshot")

    # ✅ REQUIRED CHECK
    if not user_id or not amount or not utr or not screenshot:
        return JsonResponse(
            {"status": "error", "message": "All fields required"},
            status=400
        )

    # ✅ AMOUNT VALIDATION
    try:
        amount = float(amount)
    except ValueError:
        return JsonResponse(
            {"status": "error", "message": "Invalid amount"},
            status=400
        )

    if amount <= 0:
        return JsonResponse(
            {"status": "error", "message": "Amount must be greater than zero"},
            status=400
        )

    # ✅ GET STUDENT (USER → STUDENT RELATION)
    student = Student.objects.filter(user_id=user_id).first()
    if not student:
        return JsonResponse(
            {"status": "error", "message": "Student not found"},
            status=404
        )

    # ✅ COURSE TOTAL
    total_amount = float(student.amount or 0)

    # ✅ APPROVED PAYMENTS ONLY
    paid_amount = (
        Payment.objects
        .filter(student=student, status="approved")
        .aggregate(total=Sum("amount_paid"))
        .get("total") or 0
    )
    paid_amount = float(paid_amount)

    due_amount = max(total_amount - paid_amount, 0)

    # ❌ BLOCK OVERPAYMENT
    if amount > due_amount:
        return JsonResponse(
            {
                "status": "error",
                "message": f"Amount cannot exceed due amount (₹{due_amount})"
            },
            status=400
        )

    # ❌ BLOCK DUPLICATE UTR
    if Payment.objects.filter(utr=utr).exists():
        return JsonResponse(
            {"status": "error", "message": "UTR already exists"},
            status=400
        )

    # ✅ SAVE PAYMENT (PENDING)
    Payment.objects.create(
        student=student,
        amount_paid=amount,
        utr=utr,
        screenshot=screenshot,
        status="pending"
    )

    return JsonResponse(
        {
            "status": "success",
            "message": "Payment submitted for admin approval"
        }
    )

# ==================================================
# STUDENT – PAYMENT AMOUNT INFO (API)
# ==================================================

@csrf_exempt
def payment_amount_api(request):
    user_id = request.GET.get("user_id")

    if not user_id:
        return JsonResponse(
            {"status": "error", "message": "user_id required"},
            status=400
        )

    student = Student.objects.filter(user__id=user_id).first()
    if not student:
        return JsonResponse(
            {"status": "error", "message": "Student not found"},
            status=404
        )

    total_amount = float(student.amount or 0)

    paid_amount = (
        Payment.objects
        .filter(student=student, status="approved")
        .aggregate(total=Sum("amount_paid"))
        .get("total") or 0
    )
    paid_amount = float(paid_amount)

    due_amount = max(total_amount - paid_amount, 0)

    last_payment = (
        Payment.objects
        .filter(student=student)
        .order_by("-created_at")
        .first()
    )

    return JsonResponse({
        "status": "success",
        "total_amount": total_amount,
        "paid_amount": paid_amount,
        "due_amount": due_amount,
        "payment_status": last_payment.status if last_payment else "pending",
        "admin_remark": last_payment.admin_remark if last_payment else "",
    })


# ==================================================
# ADMIN – EDIT PAYMENT
# ==================================================
@login_required
def edit_payment(request, payment_id):
    payment = get_object_or_404(Payment, id=payment_id)

    if request.method == "POST":
        payment.status = request.POST.get("status")
        payment.admin_remark = request.POST.get("admin_remark")
        payment.save()

        messages.success(request, "Payment updated successfully")

    return redirect("admin_payment_list")



from datetime import datetime

@login_required
@user_passes_test(is_admin_or_mentor)
def add_task(request):

    if request.method == "GET":
        students = Student.objects.select_related("user", "course").all()
        return render(request, "add_task.html", {"students": students})

    student = get_object_or_404(Student, id=request.POST.get("student"))

    deadline_raw = request.POST.get("deadline")
    deadline = (
        datetime.strptime(deadline_raw, "%Y-%m-%d").date()
        if deadline_raw else None
    )

    Topic.objects.create(
        content_type="task",
        student=student,
        batch=student.batch,
        mentor=getattr(request.user, "mentor", None),

        title=request.POST.get("title"),
        description=request.POST.get("task_notes") or "",

        date=timezone.localdate(),
        start_time=timezone.now().time(),
        end_time=timezone.now().time(),

        estimated_time=(
            float(request.POST.get("estimated_time"))
            if request.POST.get("estimated_time") else None
        ),

        deadline=deadline,   # ✅ FIXED
        task_notes=request.POST.get("task_notes") or "",
        status=request.POST.get("status") or "pending",
    )

    messages.success(request, "Task added successfully")
    return redirect("task_list")



@login_required
def task_list(request):

    tasks = (
        Topic.objects
        .filter(content_type="task")
        .select_related(
            "student",
            "student__user",
            "student__course",
            "student__batch"
        )
        .prefetch_related("submissions")
        .order_by("-date", "-id")
    )

    return render(request, "task_details.html", {

        "tasks": tasks
    })

from datetime import datetime

@login_required
def edit_task(request, task_id):

    task = get_object_or_404(
        Topic,
        id=task_id,
        content_type="task"
    )

    if request.method == "POST":
        deadline_raw = request.POST.get("deadline")

        task.deadline = (
            datetime.strptime(deadline_raw, "%Y-%m-%d").date()
            if deadline_raw else None
        )

        task.estimated_time = request.POST.get("estimated_time") or None
        task.task_notes = request.POST.get("task_notes") or ""
        task.status = request.POST.get("status") or "pending"

        task.save()

        messages.success(request, "Task updated successfully")

    return redirect("task_list")



@login_required
def add_topic(request):

    user = request.user
    is_admin = user.is_superuser
    mentor = Mentor.objects.filter(user=user).first()

    if request.method == "GET":
        batches = Batch.objects.select_related("course").all()
        return render(request, "add_topic.html", {
            "is_admin": is_admin,
            "batches": batches,
        })

    batch = get_object_or_404(Batch, id=request.POST.get("batch"))
    students = Student.objects.filter(batch=batch)

    for s in students:
        Topic.objects.create(
            content_type="topic",
            student=s,
            batch=batch,
            mentor=mentor,

            title=request.POST.get("title"),
            description=request.POST.get("description"),
            date=request.POST.get("date"),
            start_time=request.POST.get("start_time"),
            end_time=request.POST.get("end_time"),
            trainer=request.POST.get("trainer"),
            zoom_link=request.POST.get("zoom_link"),
            video=request.FILES.get("video"),
        )

    messages.success(request, "Topic added successfully")
    return redirect("admin_dashboard")
